from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Optional, List
import json, os, tempfile
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

app = FastAPI(title="CSR Summit 2025 CRM")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

DATA_FILE = "data.json"

def load_data():
    if not os.path.exists(DATA_FILE):
        save_data({
            "pre_registered": [], "walk_ins": [], "speakers": [],
            "counters": {"pre_individual": 0, "pre_group": 0, "walkin_individual": 0, "walkin_group": 0}
        })
    with open(DATA_FILE, "r") as f:
        return json.load(f)

def save_data(data):
    with open(DATA_FILE, "w") as f:
        json.dump(data, f, indent=2, default=str)

# ── ID helpers ──────────────────────────────────────────────────────────────
# Individual pre-reg  : PRE-0001, PRE-0002 …
# Group pre-reg       : GRP-001-01, GRP-001-02 … (same group prefix GRP-001)
# Individual walk-in  : WIN-0001, WIN-0002 …
# Group walk-in       : WGP-001-01, WGP-001-02 … (same group prefix WGP-001)

def pre_ind_id(n):  return f"PRE-{n:04d}"
def pre_grp_id(g, m): return f"GRP-{g:03d}-{m:02d}"
def win_ind_id(n):  return f"WIN-{n:04d}"
def win_grp_id(g, m): return f"WGP-{g:03d}-{m:02d}"

# ── Pydantic models ──────────────────────────────────────────────────────────
class Person(BaseModel):
    name: str; email: str = ""; phone: str = ""; organization: str = ""; designation: str = ""

class Registration(BaseModel):
    is_group: bool = False
    # individual fields (used when is_group=False)
    name: str = ""; email: str = ""; phone: str = ""; organization: str = ""; designation: str = ""
    # group fields (used when is_group=True)
    group_organization: str = ""
    group_members: Optional[List[Person]] = None

class Speaker(BaseModel):
    name: str; designation: str; organization: str; topic: str; session_time: str; bio: str = ""

# ── Stats ────────────────────────────────────────────────────────────────────
@app.get("/api/stats")
def stats():
    d = load_data()
    all_a = d["pre_registered"] + d["walk_ins"]
    ci = sum(1 for a in all_a if a.get("checked_in"))
    return {
        "pre_registered": len(d["pre_registered"]),
        "walk_ins": len(d["walk_ins"]),
        "total": len(all_a),
        "checked_in": ci,
        "pending": len(all_a) - ci,
        "speakers": len(d["speakers"])
    }

# ── Search by ID ─────────────────────────────────────────────────────────────
@app.get("/api/search")
def search(q: str = Query(...)):
    d = load_data()
    all_p = d["pre_registered"] + d["walk_ins"]
    q = q.strip().upper()
    # 1. Exact ID match
    exact = [p for p in all_p if p.get("id", "").upper() == q]
    if exact:
        return {"match_type": "individual", "query": q, "count": len(exact), "results": exact}
    # 2. Group-prefix match: GRP-001 matches GRP-001-01, GRP-001-02 …
    group_hits = [p for p in all_p if p.get("group_id", "").upper() == q]
    if group_hits:
        return {"match_type": "group", "query": q, "count": len(group_hits), "results": group_hits}
    # 3. Partial name / org search
    name_hits = [p for p in all_p if q in p.get("name", "").upper() or q in p.get("organization", "").upper()]
    if name_hits:
        return {"match_type": "name_search", "query": q, "count": len(name_hits), "results": name_hits}
    return {"match_type": "none", "query": q, "count": 0, "results": []}

# ── Pre-Registration ─────────────────────────────────────────────────────────
@app.get("/api/pre-registered")
def get_pre(): return load_data()["pre_registered"]

@app.post("/api/pre-registered")
def add_pre(reg: Registration):
    d = load_data(); now = datetime.now().isoformat()
    if reg.is_group and reg.group_members:
        d["counters"]["pre_group"] += 1
        gn = d["counters"]["pre_group"]
        gid = f"GRP-{gn:03d}"
        records = []
        for i, m in enumerate(reg.group_members, 1):
            r = {"id": pre_grp_id(gn, i), "group_id": gid, "type": "pre-registered", "is_group": True,
                 "name": m.name, "email": m.email, "phone": m.phone,
                 "organization": m.organization or reg.group_organization,
                 "designation": m.designation,
                 "checked_in": False, "check_in_time": None, "registered_at": now}
            d["pre_registered"].append(r); records.append(r)
        save_data(d)
        return {"group_id": gid, "count": len(records), "records": records}
    else:
        d["counters"]["pre_individual"] += 1
        r = {"id": pre_ind_id(d["counters"]["pre_individual"]), "group_id": None, "type": "pre-registered",
             "is_group": False, "name": reg.name, "email": reg.email, "phone": reg.phone,
             "organization": reg.organization, "designation": reg.designation,
             "checked_in": False, "check_in_time": None, "registered_at": now}
        d["pre_registered"].append(r); save_data(d)
        return r

@app.patch("/api/pre-registered/{pid}")
def toggle_pre(pid: str):
    d = load_data()
    for p in d["pre_registered"]:
        if p["id"] == pid:
            p["checked_in"] = not p["checked_in"]
            p["check_in_time"] = datetime.now().isoformat() if p["checked_in"] else None
            save_data(d); return p
    raise HTTPException(404, "Not found")

@app.delete("/api/pre-registered/{pid}")
def del_pre(pid: str):
    d = load_data()
    d["pre_registered"] = [p for p in d["pre_registered"] if p["id"] != pid]
    save_data(d); return {"ok": True}

@app.post("/api/pre-registered/bulk")
def bulk_import(records: List[dict]):
    d = load_data(); now = datetime.now().isoformat(); added = []
    for rec in records:
        d["counters"]["pre_individual"] += 1
        r = {"id": pre_ind_id(d["counters"]["pre_individual"]), "group_id": None, "type": "pre-registered",
             "is_group": False, "name": rec.get("name",""), "email": rec.get("email",""),
             "phone": rec.get("phone",""), "organization": rec.get("organization",""),
             "designation": rec.get("designation",""),
             "checked_in": False, "check_in_time": None, "registered_at": now}
        d["pre_registered"].append(r); added.append(r)
    save_data(d); return {"added": len(added), "records": added}

# ── Walk-ins ─────────────────────────────────────────────────────────────────
@app.get("/api/walk-ins")
def get_walkins(): return load_data()["walk_ins"]

@app.post("/api/walk-ins")
def add_walkin(reg: Registration):
    d = load_data(); now = datetime.now().isoformat()
    if reg.is_group and reg.group_members:
        d["counters"]["walkin_group"] += 1
        gn = d["counters"]["walkin_group"]
        gid = f"WGP-{gn:03d}"
        records = []
        for i, m in enumerate(reg.group_members, 1):
            r = {"id": win_grp_id(gn, i), "group_id": gid, "type": "walk-in", "is_group": True,
                 "name": m.name, "email": m.email, "phone": m.phone,
                 "organization": m.organization or reg.group_organization,
                 "designation": m.designation,
                 "checked_in": True, "check_in_time": now, "registered_at": now}
            d["walk_ins"].append(r); records.append(r)
        save_data(d)
        return {"group_id": gid, "count": len(records), "records": records}
    else:
        d["counters"]["walkin_individual"] += 1
        r = {"id": win_ind_id(d["counters"]["walkin_individual"]), "group_id": None, "type": "walk-in",
             "is_group": False, "name": reg.name, "email": reg.email, "phone": reg.phone,
             "organization": reg.organization, "designation": reg.designation,
             "checked_in": True, "check_in_time": now, "registered_at": now}
        d["walk_ins"].append(r); save_data(d)
        return r

@app.patch("/api/walk-ins/{wid}")
def toggle_walkin(wid: str):
    d = load_data()
    for p in d["walk_ins"]:
        if p["id"] == wid:
            p["checked_in"] = not p["checked_in"]
            p["check_in_time"] = datetime.now().isoformat() if p["checked_in"] else None
            save_data(d); return p
    raise HTTPException(404, "Not found")

@app.delete("/api/walk-ins/{wid}")
def del_walkin(wid: str):
    d = load_data()
    d["walk_ins"] = [p for p in d["walk_ins"] if p["id"] != wid]
    save_data(d); return {"ok": True}

# ── Speakers ─────────────────────────────────────────────────────────────────
@app.get("/api/speakers")
def get_speakers(): return load_data()["speakers"]

@app.post("/api/speakers")
def add_speaker(s: Speaker):
    d = load_data()
    r = {"id": f"SPK-{len(d['speakers'])+1:03d}", **s.dict(), "added_at": datetime.now().isoformat()}
    d["speakers"].append(r); save_data(d); return r

@app.delete("/api/speakers/{sid}")
def del_speaker(sid: str):
    d = load_data()
    d["speakers"] = [s for s in d["speakers"] if s["id"] != sid]
    save_data(d); return {"ok": True}

# ── Excel Export ─────────────────────────────────────────────────────────────
def hdr(ws, headers, fill="1e3a5f"):
    f = PatternFill("solid", fgColor=fill)
    fw = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = f; cell.font = fw
        cell.alignment = Alignment(horizontal="center")

def attendee_sheet(ws, records):
    hdr(ws, ["ID","Group ID","Name","Email","Phone","Organization","Designation","Type","Group?","Status","Check-in Time","Registered At"])
    for r in records:
        ws.append([r.get("id"), r.get("group_id") or "", r.get("name"), r.get("email"),
                   r.get("phone"), r.get("organization"), r.get("designation"), r.get("type"),
                   "Yes" if r.get("is_group") else "No",
                   "Present" if r.get("checked_in") else "Absent",
                   r.get("check_in_time") or "", r.get("registered_at") or ""])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

@app.get("/api/export/excel")
def export(type: str = "all"):
    d = load_data()
    pre = d["pre_registered"]; walkins = d["walk_ins"]
    all_a = pre + walkins; speakers = d["speakers"]
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    if type in ("all", "attendees"):
        ws = wb.create_sheet("All Attendees"); attendee_sheet(ws, all_a)
    if type in ("all", "preregistered"):
        ws = wb.create_sheet("Pre-Registered"); attendee_sheet(ws, pre)
    if type in ("all", "walkins"):
        ws = wb.create_sheet("Walk-Ins"); attendee_sheet(ws, walkins)
    if type in ("all", "attendance"):
        ws = wb.create_sheet("Attendance")
        hdr(ws, ["ID","Group ID","Name","Organization","Designation","Type","Status","Check-in Time"])
        for r in all_a:
            ws.append([r.get("id"), r.get("group_id") or "", r.get("name"), r.get("organization"),
                       r.get("designation"), r.get("type"),
                       "Present" if r.get("checked_in") else "Absent", r.get("check_in_time") or ""])
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 20
    if type in ("all", "speakers"):
        ws = wb.create_sheet("Speakers")
        hdr(ws, ["ID","Name","Designation","Organization","Topic","Session Time","Bio"])
        for s in speakers:
            ws.append([s.get("id"), s.get("name"), s.get("designation"), s.get("organization"),
                       s.get("topic"), s.get("session_time"), s.get("bio")])
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 22
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name); tmp.close()
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    return FileResponse(tmp.name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        filename=f"CSR_Summit_{type}_{ts}.xlsx")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
