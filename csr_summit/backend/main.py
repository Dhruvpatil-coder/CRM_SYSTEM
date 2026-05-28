"""
CSR Summit 2025 - Backend API
FastAPI + Excel Export
Run: uvicorn main:app --reload --port 8000
"""

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Optional, List
import json
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile

app = FastAPI(title="CSR Summit CRM API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─── In-memory store (replace with DB for production) ───
DB_FILE = "data.json"

def load_db():
    if os.path.exists(DB_FILE):
        with open(DB_FILE) as f:
            return json.load(f)
    return {"pre_registered": [], "walk_ins": [], "speakers": [], "attendance": {}}

def save_db(data):
    with open(DB_FILE, "w") as f:
        json.dump(data, f, indent=2, default=str)

db = load_db()

# ─── Models ───
class Attendee(BaseModel):
    name: str
    email: str
    phone: Optional[str] = ""
    organization: Optional[str] = ""
    designation: Optional[str] = ""

class AttendeeUpdate(BaseModel):
    checked_in: bool

class Speaker(BaseModel):
    name: str
    designation: str
    organization: Optional[str] = ""
    topic: str
    session: Optional[str] = ""
    bio: Optional[str] = ""



def make_id(prefix, items):
    return f"{prefix}-{str(len(items)+1).zfill(3)}"

def sync_to_sheets(spreadsheet_id: str, creds_dict: dict):
    """Sync all data to Google Sheets."""
    if not SHEETS_AVAILABLE:
        return {"error": "gspread not installed. Run: pip install gspread google-auth"}
    
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = Credentials.from_service_account_info(creds_dict, scopes=scope)
    client = gspread.authorize(creds)
    sh = client.open_by_key(spreadsheet_id)

    # Attendees sheet
    all_attendees = [
        {**a, "type": "pre-registered"} for a in db["pre_registered"]
    ] + [
        {**a, "type": "walk-in"} for a in db["walk_ins"]
    ]
    attendee_rows = [["ID","Name","Email","Phone","Organization","Designation","Type","Checked In","Registered At"]]
    for a in all_attendees:
        attendee_rows.append([
            a.get("id",""), a.get("name",""), a.get("email",""),
            a.get("phone",""), a.get("organization",""), a.get("designation",""),
            a.get("type",""), "Yes" if a.get("checked_in") else "No",
            a.get("registered_at","")
        ])
    try:
        ws = sh.worksheet("Attendees")
        ws.clear()
    except:
        ws = sh.add_worksheet("Attendees", rows=500, cols=10)
    ws.update("A1", attendee_rows)

    # Speakers sheet
    speaker_rows = [["ID","Name","Designation","Organization","Topic","Session","Bio"]]
    for s in db["speakers"]:
        speaker_rows.append([s.get("id",""), s.get("name",""), s.get("designation",""),
            s.get("organization",""), s.get("topic",""), s.get("session",""), s.get("bio","")])
    try:
        ws2 = sh.worksheet("Speakers")
        ws2.clear()
    except:
        ws2 = sh.add_worksheet("Speakers", rows=100, cols=8)
    ws2.update("A1", speaker_rows)

    return {"synced": len(all_attendees), "speakers": len(db["speakers"])}

# ─── Routes: Pre-Registration ───
@app.get("/api/pre-registered")
def get_pre_registered():
    return db["pre_registered"]

@app.post("/api/pre-registered", status_code=201)
def add_pre_registered(attendee: Attendee):
    record = {
        "id": make_id("PRE", db["pre_registered"]),
        **attendee.dict(),
        "checked_in": False,
        "type": "pre-registered",
        "registered_at": datetime.now().strftime("%d/%m/%Y")
    }
    db["pre_registered"].append(record)
    save_db(db)
    return record

@app.patch("/api/pre-registered/{attendee_id}")
def update_pre_registered(attendee_id: str, update: AttendeeUpdate):
    for a in db["pre_registered"]:
        if a["id"] == attendee_id:
            a["checked_in"] = update.checked_in
            save_db(db)
            return a
    raise HTTPException(404, "Attendee not found")

@app.delete("/api/pre-registered/{attendee_id}")
def delete_pre_registered(attendee_id: str):
    db["pre_registered"] = [a for a in db["pre_registered"] if a["id"] != attendee_id]
    save_db(db)
    return {"deleted": attendee_id}

# ─── Routes: Walk-in ───
@app.get("/api/walk-ins")
def get_walk_ins():
    return db["walk_ins"]

@app.post("/api/walk-ins", status_code=201)
def add_walk_in(attendee: Attendee):
    record = {
        "id": make_id("WI", db["walk_ins"]),
        **attendee.dict(),
        "checked_in": True,  # walk-ins auto check-in
        "type": "walk-in",
        "registered_at": datetime.now().strftime("%d/%m/%Y %H:%M")
    }
    db["walk_ins"].append(record)
    save_db(db)
    return record

@app.patch("/api/walk-ins/{attendee_id}")
def update_walk_in(attendee_id: str, update: AttendeeUpdate):
    for a in db["walk_ins"]:
        if a["id"] == attendee_id:
            a["checked_in"] = update.checked_in
            save_db(db)
            return a
    raise HTTPException(404, "Attendee not found")

# ─── Routes: Speakers ───
@app.get("/api/speakers")
def get_speakers():
    return db["speakers"]

@app.post("/api/speakers", status_code=201)
def add_speaker(speaker: Speaker):
    initials = "".join(w[0] for w in speaker.name.split()[:2]).upper()
    colors = ["#1565c0","#00897b","#7b1fa2","#e65100","#c62828","#2e7d32","#ad1457"]
    color = colors[len(db["speakers"]) % len(colors)]
    record = {
        "id": make_id("SP", db["speakers"]),
        **speaker.dict(),
        "avatar": initials,
        "color": color
    }
    db["speakers"].append(record)
    save_db(db)
    return record

@app.delete("/api/speakers/{speaker_id}")
def delete_speaker(speaker_id: str):
    db["speakers"] = [s for s in db["speakers"] if s["id"] != speaker_id]
    save_db(db)
    return {"deleted": speaker_id}

# ─── Routes: Stats ───
@app.get("/api/stats")
def get_stats():
    all_attendees = db["pre_registered"] + db["walk_ins"]
    checked_in = sum(1 for a in all_attendees if a.get("checked_in"))
    return {
        "pre_registered": len(db["pre_registered"]),
        "walk_ins": len(db["walk_ins"]),
        "total": len(all_attendees),
        "checked_in": checked_in,
        "pending": len(all_attendees) - checked_in,
        "speakers": len(db["speakers"])
    }

# ─── Routes: Export ───
@app.get("/api/export/excel")
def export_excel(type: str = "all"):
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="0F2D4A")
    center = Alignment(horizontal="center", vertical="center")

    def style_header(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

    all_att = (
        [{**a, "type": "pre-registered"} for a in db["pre_registered"]] +
        [{**a, "type": "walk-in"} for a in db["walk_ins"]]
    )

    def add_attendees_sheet(ws, rows):
        style_header(ws, ["ID","Name","Email","Phone","Organization","Designation","Type","Checked In","Registered At"])
        for a in rows:
            ws.append([a.get("id"),a.get("name"),a.get("email"),a.get("phone"),
                       a.get("organization"),a.get("designation"),a.get("type",""),
                       "Yes" if a.get("checked_in") else "No",a.get("registered_at")])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or ""))+4 for c in col)

    first = True

    if type in ("all", "attendees"):
        ws = wb.active if first else wb.create_sheet("All Attendees")
        ws.title = "All Attendees"; first = False
        add_attendees_sheet(ws, all_att)

    if type in ("all", "preregistered"):
        ws = wb.active if first else wb.create_sheet("Pre-Registered")
        ws.title = "Pre-Registered"; first = False
        add_attendees_sheet(ws, [{**a,"type":"pre-registered"} for a in db["pre_registered"]])

    if type in ("all", "walkins"):
        ws = wb.active if first else wb.create_sheet("Walk-Ins")
        ws.title = "Walk-Ins"; first = False
        style_header(ws, ["ID","Name","Email","Phone","Organization","Designation","Registered At"])
        for a in db["walk_ins"]:
            ws.append([a.get("id"),a.get("name"),a.get("email"),a.get("phone"),
                       a.get("organization"),a.get("designation"),a.get("registered_at")])

    if type in ("all", "attendance"):
        ws = wb.active if first else wb.create_sheet("Attendance")
        ws.title = "Attendance"; first = False
        style_header(ws, ["ID","Name","Organization","Type","Status","Time"])
        for a in all_att:
            ws.append([a.get("id"),a.get("name"),a.get("organization"),
                       a.get("type"),"Present" if a.get("checked_in") else "Absent",
                       a.get("registered_at","")])

    if type in ("all", "speakers"):
        ws = wb.active if first else wb.create_sheet("Speakers")
        ws.title = "Speakers"
        style_header(ws, ["ID","Name","Designation","Organization","Topic","Session","Bio"])
        for s in db["speakers"]:
            ws.append([s.get("id"),s.get("name"),s.get("designation"),s.get("organization"),
                       s.get("topic"),s.get("session"),s.get("bio")])

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    filename = f"CSR_Summit_June5_{type}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return FileResponse(tmp.name, filename=filename,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─── Bulk import pre-registrations ───
@app.post("/api/pre-registered/bulk")
def bulk_import(attendees: List[Attendee]):
    added = []
    for attendee in attendees:
        record = {
            "id": make_id("PRE", db["pre_registered"]),
            **attendee.dict(),
            "checked_in": False,
            "type": "pre-registered",
            "registered_at": datetime.now().strftime("%d/%m/%Y")
        }
        db["pre_registered"].append(record)
        added.append(record)
    save_db(db)
    return {"added": len(added), "records": added}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
