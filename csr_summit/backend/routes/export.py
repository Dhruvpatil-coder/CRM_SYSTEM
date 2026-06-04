# ── Excel Export ─────────────────────────────────────────────────────────────
import tempfile
import openpyxl

from fastapi import APIRouter
from fastapi.responses import FileResponse
from datetime import datetime

from openpyxl.styles import PatternFill, Font, Alignment

from supabase_client import supabase

router = APIRouter()


def hdr(ws, headers, fill="1e3a5f"):
    f = PatternFill("solid", fgColor=fill)
    fw = Font(bold=True, color="FFFFFF")

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = f
        cell.font = fw
        cell.alignment = Alignment(horizontal="center")


def attendee_sheet(ws, records):
    hdr(
        ws,
        [
            "Name",
            "Email",
            "Phone",
            "Organization",
            "Designation",
            "Type",
            "Group?",
            "Status",
            "Check-in Time",
        ],
    )

    for r in records:
        ws.append(
            [
                r.get("name"),
                r.get("email"),
                r.get("phone"),
                r.get("organization"),
                r.get("designation"),
                r.get("type"),
                "Yes" if r.get("is_group") else "No",
                "Present" if r.get("checked_in") else "Absent",
                r.get("checked_in_time") or "",
            ]
        )

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20


@router.get("/api/export/excel")
def export(type: str = "all"):

    attendees = (
        supabase
        .table("attendees")
        .select("*")
        .execute()
        .data
    )

    speakers = (
        supabase
        .table("speakers")
        .select("*")
        .execute()
        .data
    )

    pre = [
        a for a in attendees
        if a.get("type") == "pre-registered"
    ]

    walkins = [
        a for a in attendees
        if a.get("type") == "walk-in"
    ]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if type in ("all", "attendees"):
        ws = wb.create_sheet("All Attendees")
        attendee_sheet(ws, attendees)

    if type in ("all", "preregistered"):
        ws = wb.create_sheet("Pre-Registered")
        attendee_sheet(ws, pre)

    if type in ("all", "walkins"):
        ws = wb.create_sheet("Walk-Ins")
        attendee_sheet(ws, walkins)

    if type in ("all", "attendance"):
        ws = wb.create_sheet("Attendance")

        hdr(
            ws,
            [
                "Phone",
                "Name",
                "Organization",
                "Designation",
                "Type",
                "Status",
                "Check-in Time",
            ],
        )

        for r in attendees:
            ws.append(
                [
                    r.get("phone"),
                    r.get("name"),
                    r.get("organization"),
                    r.get("designation"),
                    r.get("type"),
                    "Present" if r.get("checked_in") else "Absent",
                    r.get("checked_in_time") or "",
                ]
            )

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20

    if type in ("all", "speakers"):
        ws = wb.create_sheet("Speakers")

        hdr(
            ws,
            [
                "Name",
                "Organization",
                "Topic",
                "Session Time",
                "Bio",
            ],
        )

        for s in speakers:
            ws.append(
                [
                    s.get("name"),
                    s.get("organization"),
                    s.get("topic"),
                    s.get("session_time"),
                    s.get("bio"),
                ]
            )

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 22

    tmp = tempfile.NamedTemporaryFile(
        suffix=".xlsx",
        delete=False
    )

    wb.save(tmp.name)
    tmp.close()

    ts = datetime.now().strftime("%Y%m%d_%H%M")

    return FileResponse(
        tmp.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"CSR_Summit_{type}_{ts}.xlsx"
    )
